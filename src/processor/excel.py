from openpyxl import load_workbook
from datetime import datetime
import jsonpickle
import json
import os


class SheetData:
        def __init__(self, name, rows):
                self.name = name
                self.rows = rows

class SheetDataEncoder(json.JSONEncoder):
        def default(self, obj):
                if isinstance(obj, SheetData):
                        return [obj.name, obj.rows]
                if isinstance(obj, datetime):
                        return str(obj)
                return json.JSONEncoder.default(self, obj)

def load_file(filename):
        try:
                workbook = load_workbook(filename = filename)
                sheet_names = workbook.sheetnames

                sheet_list = []
                for sheet_name in sheet_names:
                        print(sheet_name)
                        sheet = workbook[sheet_name]

                        row_list = []
                        for row in sheet.rows:
                                cell_list = []
                                for cell in row:
                                        cell_list.append(str(cell.value))
                                row_list.append(cell_list)

                        sheet_data = SheetData(sheet_name, row_list)
                        sheet_list.append(sheet_data)

                print(json.dumps(sheet_list, cls=SheetDataEncoder, indent=4))
                return sheet_list
        except:
                print("Error in reading the file")
                return []

def store_as_file(filename, sheet_list, output_folder):
        ensure_output(output_folder)

        for sheet in sheet_list:
                target_file = get_csv_file_name(filename, sheet.name)
                with open(f"{output_folder}\\{target_file}", "w") as target_file:
                        data = list(map(lambda x : ",".join(x), sheet.rows))
                        target_file.write("\n".join(data))

def get_csv_file_name(filename, target):
        path, file = os.path.split(filename)
        name, extension = os.path.splitext(file)

        return f"{name}_{target}.csv"

def ensure_output(output_folder):
        if not os.path.exists(output_folder):
                os.makedirs(output_folder)








        

                

        
    

